import csv
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook

INPUT_FILE = "Receiving.csv"
OUTPUT_WORKBOOK = "Receiving_Report.xlsx"

# Column indexes (0-based)
COL_TRUCK = 0          # A
COL_TICKET = 6         # G
COL_RAILCAR = 7        # H
COL_MATERIAL_LOT = 8   # I
COL_PILE = 9           # J
COL_REMARKS = 11       # L
COL_DATE = 16          # Q
COL_GROSS = 20         # U
COL_TARE = 21          # V
COL_NET = 22           # W


def parse_date(value):
    try:
        return datetime.strptime(value.strip(), "%m/%d/%Y").date()
    except Exception:
        return None


def parse_weight(value):
    try:
        return float(value.replace(",", "").strip())
    except Exception:
        return None


daily_summary = defaultdict(lambda: {"trucks": 0, "weight": 0.0})
daily_pile_lot_summary = defaultdict(float)
detail_rows = []

with open(INPUT_FILE, newline="", encoding="utf-8") as infile:
    reader = csv.reader(infile)

    for row in reader:
        if len(row) <= COL_NET:
            continue

        # ---------------- FILTERS ----------------
        if not row[COL_RAILCAR].strip():
            continue  # remove railcars

        if not row[COL_PILE].strip():
            continue  # remove voided tickets

        # -----------------------------------------
        date = parse_date(row[COL_DATE])
        net_weight = parse_weight(row[COL_NET])

        if date is None or net_weight is None:
            continue

        truck = row[COL_TRUCK]
        ticket = row[COL_TICKET]
        pile = row[COL_PILE]
        material_lot = row[COL_MATERIAL_LOT]
        remarks = row[COL_REMARKS]

        gross = parse_weight(row[COL_GROSS])
        tare = parse_weight(row[COL_TARE])

        # -------- Receiving1 --------
        daily_summary[date]["trucks"] += 1
        daily_summary[date]["weight"] += net_weight

        # -------- Receiving2 --------
        daily_pile_lot_summary[(date, pile, material_lot)] += net_weight

        # -------- Receiving3 --------
        detail_rows.append([
            truck,
            ticket,
            pile,
            date.strftime("%m/%d/%Y"),
            material_lot,
            remarks,
            int(gross) if gross is not None else "",
            int(tare) if tare is not None else "",
            int(net_weight)
        ])


# ================= Workbook =================

wb = Workbook()

# ---- Receiving1 ----
ws_daily = wb.active
ws_daily.title = "Receiving1"
ws_daily.append(["Date", "Total Trucks Received", "Total Net Weight Received"])

for date in sorted(daily_summary):
    ws_daily.append([
        date.strftime("%m/%d/%Y"),
        daily_summary[date]["trucks"],
        int(daily_summary[date]["weight"])
    ])


# ---- Receiving2 ----
ws_pile_lot = wb.create_sheet("Receiving2")
ws_pile_lot.append(["Date", "Pile Number", "Material Lot #", "Total Weight"])

for (date, pile, material_lot), weight in sorted(daily_pile_lot_summary.items()):
    ws_pile_lot.append([
        date.strftime("%m/%d/%Y"),
        pile,
        material_lot,
        int(weight)
    ])


# ---- Receiving3 ----
ws_details = wb.create_sheet("Receiving3")
ws_details.append([
    "Truck #",
    "Ticket ID",
    "Pile #",
    "Date",
    "Material Lot #",
    "Remarks",
    "Gross Weight",
    "Tare Weight",
    "Net Weight"
])

for row in detail_rows:
    ws_details.append(row)

wb.save(OUTPUT_WORKBOOK)

print(f"Workbook created: {OUTPUT_WORKBOOK}")
print(f"Filtered rows written: {len(detail_rows):,}")
