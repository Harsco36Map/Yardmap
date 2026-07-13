#!/usr/bin/env python3
from pathlib import Path
import csv
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook

DEFAULT_LBS_PER_TON = 2000


# -------------------------------------------------
# Utilities
# -------------------------------------------------

def detect_delimiter(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        sample = f.read(4096)
    try:
        return csv.Sniffer().sniff(sample).delimiter
    except Exception:
        return ","


def parse_date(val: str):
    if not val:
        return None
    val = val.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            pass
    return None


def to_float(val: str):
    if not val:
        return None
    try:
        return float(val.replace(",", ""))
    except ValueError:
        return None


def get_heat_number(heat_bucket: str) -> str:
    return heat_bucket.split("/")[0].strip()


def get_heat_type(raw_heat_type: str) -> str:
    return raw_heat_type.strip()


# -------------------------------------------------
# Main Processing
# -------------------------------------------------

def main():
    # Use the most recent CSV in the current directory
    src = next(Path.cwd().glob("*.csv"))
    sep = detect_delimiter(src)

    # ---- Primary accumulators ----
    daily_totals = defaultdict(float)

    pile_daily = defaultdict(lambda: defaultdict(float))
    pile_grand = defaultdict(float)
    pile_daily_lots = defaultdict(lambda: defaultdict(set))
    pile_grand_lots = defaultdict(set)

    completed_heat_rows = []
    heat_types = {}

    # ---- Bucket / Heat tracking ----
    daily_buckets = defaultdict(set)
    daily_heats = defaultdict(set)

    bucket_last_date = {}
    heat_last_date = {}

    # -------------------------------------------------
    # FIRST PASS — determine FINAL DATE for each bucket & heat
    # -------------------------------------------------

    rows = []
    with src.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f, delimiter=sep)
        for row in reader:
            rows.append(row)

            if len(row) < 11:
                continue

            heat_bucket = row[0].strip()
            raw_heat_type = row[1].strip() if len(row) > 1 else ""
            date = parse_date(row[10])

            if not heat_bucket or date is None:
                continue

            heat_only = get_heat_number(heat_bucket)

            if raw_heat_type and heat_only not in heat_types:
                heat_types[heat_only] = get_heat_type(raw_heat_type)

            if heat_bucket not in bucket_last_date or date > bucket_last_date[heat_bucket]:
                bucket_last_date[heat_bucket] = date

            if heat_only not in heat_last_date or date > heat_last_date[heat_only]:
                heat_last_date[heat_only] = date

    # -------------------------------------------------
    # SECOND PASS — aggregate data with final‑day logic
    # -------------------------------------------------

    for row in rows:
        if len(row) < 11:
            continue

        heat_bucket = row[0].strip()
        pile = row[6].strip()           # Pile Number (Column G)
        material = row[7].strip()       # Material Lot # (reference only)
        consumed = to_float(row[9])     # Weight consumed
        date = parse_date(row[10])      # Date loaded

        if not heat_bucket or consumed is None or date is None:
            continue

        heat_only = get_heat_number(heat_bucket)
        completion_date = heat_last_date.get(heat_only)

        # ---- Daily total consumed (always counted on actual load date) ----
        daily_totals[date] += consumed

        if pile:
            pile_daily[date][pile] += consumed
            pile_grand[pile] += consumed
            if material:
                pile_daily_lots[date][pile].add(material)
                pile_grand_lots[pile].add(material)

        # ---- Bucket / Heat counts (ONLY on final date) ----
        if bucket_last_date.get(heat_bucket) == date:
            daily_buckets[date].add(heat_bucket)

        if heat_last_date.get(heat_only) == date:
            daily_heats[date].add(heat_only)

        if completion_date and pile:
            bucket_number = row[3].strip() if len(row) > 3 else ""
            completed_heat_rows.append({
                "date": completion_date,
                "original_heat_bucket": heat_bucket,
                "heat_number": heat_only,
                "bucket_number": bucket_number,
                "heat_type": heat_types.get(heat_only, ""),
                "pile_number": pile,
                "material_lot": material,
                "total_lbs": consumed,
                "total_tons": consumed / DEFAULT_LBS_PER_TON,
            })

    # -------------------------------------------------
    # OUTPUT — Single XLSX with 4 tabs
    # -------------------------------------------------

    wb = Workbook()

    # Sheet 1: Daily Summary
    ws_daily = wb.active
    ws_daily.title = "Consumption1"
    ws_daily.append([
        "Date",
        "Total_Lbs",
        "Total_Tons",
        "Buckets_Loaded",
        "Heats_Loaded"
    ])
    for d in sorted(daily_totals):
        lbs = daily_totals[d]
        ws_daily.append([
            d.isoformat(),
            round(lbs, 0),
            round(lbs / DEFAULT_LBS_PER_TON, 2),
            len(daily_buckets[d]),
            len(daily_heats[d])
        ])

    # Sheet 2: Daily Pile Totals (with material lot reference)
    ws_daily_pile = wb.create_sheet("Consumption2")
    ws_daily_pile.append(["Date", "Pile_Number", "Material_Lots", "Total_Lbs", "Total_Tons"])
    for d in sorted(pile_daily):
        for pile, lbs in sorted(pile_daily[d].items()):
            lots = sorted(pile_daily_lots[d][pile]) if pile in pile_daily_lots[d] else []
            ws_daily_pile.append([
                d.isoformat(),
                pile,
                "; ".join(lots),
                round(lbs, 0),
                round(lbs / DEFAULT_LBS_PER_TON, 2)
            ])

    # Sheet 3: Grand Pile Totals (with material lot reference)
    ws_grand_pile = wb.create_sheet("Consumption3")
    ws_grand_pile.append(["Pile_Number", "Material_Lots", "Total_Lbs", "Total_Tons"])
    for pile, lbs in sorted(pile_grand.items()):
        lots = sorted(pile_grand_lots[pile]) if pile in pile_grand_lots else []
        ws_grand_pile.append([
            pile,
            "; ".join(lots),
            round(lbs, 0),
            round(lbs / DEFAULT_LBS_PER_TON, 2)
        ])

    # Sheet 4: Completed heat details by completion date
    ws_completed_heats = wb.create_sheet("Consumption4")
    ws_completed_heats.append([
        "Date",
        "Original_Heat_Bucket",
        "Heat_Number",
        "Bucket_Number",
        "Heat_Type",
        "Pile_Number",
        "Material_Lots",
        "Total_Lbs",
        "Total_Tons"
    ])
    for entry in sorted(
        completed_heat_rows,
        key=lambda item: (
            item["date"],
            item["heat_number"],
            item["pile_number"],
            item["material_lot"],
            item["bucket_number"],
        ),
    ):
        ws_completed_heats.append([
            entry["date"].isoformat(),
            entry["original_heat_bucket"],
            entry["heat_number"],
            entry["bucket_number"],
            entry["heat_type"],
            entry["pile_number"],
            entry["material_lot"],
            round(entry["total_lbs"], 0),
            round(entry["total_tons"], 2),
        ])

    out_xlsx = src.with_name(f"{src.stem}_loading_reports.xlsx")
    wb.save(out_xlsx)

    print("Report generated:")
    print(f" - {out_xlsx.name}")


if __name__ == "__main__":
    main()