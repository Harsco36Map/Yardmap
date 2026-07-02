import argparse
import calendar
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

SUPPLIER_TRANSLATIONS = {
    "ARVIN SANGO INC": "Kroot",
    "PRECISION STRIP": "Kroot",
    "FAURECIA EMISSIONS": "Kroot",
    "AK TUBE LLC": "Kroot",
    "OMNISOURCE CORPO": "Omnisource",
    "FERROUS PROCESSI": "Ferrous Processing",
    "TENNECO INC": "Ferrous Processing",
    "CMC COMMERCIAL M": "Commercial Metals",
    "COMBINED METAL I": "Combined Metal",
    "CLEVELAND CLIFFS": "Rockport",
    "COHEN BROTHERS I": "Middletown",
}

HEADER_RENAMES = {
    "Scale W. Gross": "Scale Gross",
    "Scale W Tare": "Scale Tare",
}

EXTRA_COLUMNS = ["Offloaded", "Car Status"]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def ordinal(day: int) -> str:
    if 10 <= day % 100 <= 20:
        return f"{day}th"
    return f"{day}{['th','st','nd','rd','th','th','th','th','th','th'][day % 10]}"


def make_report_title(file_stem: str, year: int) -> str:
    base = re.sub(r"\s*RAW$", "", file_stem, flags=re.I).strip()
    match = re.search(r"ASNs\s*([0-9]{1,2})[-_ ]([0-9]{1,2})", base, flags=re.I)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        month_name = calendar.month_name[month]
        return f"ASNs Arrived Report {month_name} {ordinal(day)}, {year}"
    return base


def normalize_header_row(raw_headers):
    headers = []
    for heading in raw_headers:
        heading = heading.strip()
        if heading in HEADER_RENAMES:
            heading = HEADER_RENAMES[heading]
        headers.append(heading)
    return headers


def translate_supplier(name: str) -> str | None:
    if not name:
        return None
    key = name.strip().upper()
    return SUPPLIER_TRANSLATIONS.get(key)


def trim_trailing_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    num_cols = len(rows[0])
    while num_cols > 0 and not str(rows[0][num_cols - 1]).strip() and all(
        len(row) < num_cols or not str(row[num_cols - 1]).strip() for row in rows[1:]
    ):
        num_cols -= 1

    trimmed = []
    for row in rows:
        row_values = list(row)
        if len(row_values) < num_cols:
            row_values += [""] * (num_cols - len(row_values))
        trimmed.append(row_values[:num_cols])
    return trimmed


def make_border(*, top=False, bottom=False, left=False, right=False) -> Border:
    return Border(
        left=Side(style="thin") if left else Side(style=None),
        right=Side(style="thin") if right else Side(style=None),
        top=Side(style="thin") if top else Side(style=None),
        bottom=Side(style="thin") if bottom else Side(style=None),
    )


def parse_number(value: str):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return value


def is_summary_row(raw_values: list[str]) -> bool:
    if not raw_values:
        return False
    cleaned = [str(value).strip() for value in raw_values]
    if len(cleaned) < 8:
        return False
    if cleaned[7] == "" or any(cleaned[i] for i in range(len(cleaned)) if i != 7):
        return False
    return True


def build_workbook(csv_path: Path, output_path: Path) -> None:
    with csv_path.open(newline="", encoding="utf-8-sig") as csvfile:
        reader = csv.reader(csvfile)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV file is empty: {csv_path}")

    rows = trim_trailing_empty_columns(rows)
    raw_headers = rows[0]
    headers = normalize_header_row(raw_headers) + EXTRA_COLUMNS
    num_data_columns = len(raw_headers)
    num_columns = len(headers)
    worksheet_name = output_path.stem[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name

    title = make_report_title(csv_path.stem, datetime.now().year)
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, heading in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=heading)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        align = "left" if col_idx <= 5 else "center"
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    output_row = 3
    supplier_column_index = 3
    for raw_row in rows[1:]:
        raw_values = list(raw_row)
        if len(raw_values) < num_data_columns:
            raw_values += [""] * (num_data_columns - len(raw_values))
        raw_values = raw_values[:num_data_columns]

        if is_summary_row(raw_values):
            for col_idx, value in enumerate(raw_values, start=1):
                parsed_value = parse_number(value)
                cell = ws.cell(row=output_row, column=col_idx, value=parsed_value if parsed_value is not None else value)
                if col_idx == 8:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "#,##0"
                elif col_idx <= 5:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            output_row += 1
            continue

        for col_idx, value in enumerate(raw_values, start=1):
            parsed_value = parse_number(value)
            cell = ws.cell(row=output_row, column=col_idx, value=parsed_value if parsed_value is not None else value)
            if col_idx <= 5:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_idx in (2, 4):
                cell.number_format = "0"
            elif col_idx in (6, 7, 8, 9, 10):
                cell.number_format = "#,##0"

        for extra_index in range(len(EXTRA_COLUMNS)):
            col_idx = num_data_columns + extra_index + 1
            cell = ws.cell(row=output_row, column=col_idx, value=None)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        supplier_name = ""
        if len(raw_values) >= supplier_column_index:
            supplier_name = raw_values[supplier_column_index - 1]
        translation = translate_supplier(supplier_name)

        if translation:
            next_row = output_row + 1
            supplier_cell = ws.cell(row=next_row, column=supplier_column_index, value=translation)
            supplier_cell.font = Font(bold=True)
            supplier_cell.alignment = Alignment(vertical="center")
            block_rows = [output_row, next_row]
        else:
            block_rows = [output_row]

        for row_idx in block_rows:
            for col_idx in range(1, num_columns + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                top = row_idx == block_rows[0]
                bottom = row_idx == block_rows[-1]
                left = col_idx == 1
                right = col_idx == num_columns
                cell.border = make_border(top=top, bottom=bottom, left=left, right=right)

        output_row += len(block_rows)

    # Auto-fit columns to content width, excluding the merged title row
    max_widths: dict[int, int] = {col_idx: 0 for col_idx in range(1, num_columns + 1)}
    for row_idx in range(2, output_row):
        for col_idx in range(1, num_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if cell.number_format in ("#,##0", "0") and isinstance(value, (int, float)):
                length = max(length, len(str(value)))
            max_widths[col_idx] = max(max_widths[col_idx], length)

    for col_idx, width in max_widths.items():
        adjusted = min(max(width + 2, 8), 40)
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = adjusted

    wb.save(output_path)


def find_raw_csv_files(directory: Path):
    raw_files = sorted(directory.glob("* RAW.csv"))
    if raw_files:
        return raw_files
    return sorted(directory.glob("*.csv"))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert raw railcar CSV files into formatted Excel worksheets."
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="Raw CSV file to convert. If omitted, all '* RAW.csv' files in the current directory are processed.",
    )
    parser.add_argument(
        "--output-dir",
        help="Optional directory to write output worksheets into.",
        default=".",
    )
    args = parser.parse_args()

    cwd = Path.cwd()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    input_paths = []
    if args.input:
        input_paths = [Path(args.input)]
    else:
        input_paths = find_raw_csv_files(cwd)

    if not input_paths:
        raise SystemExit("No CSV files found to process.")

    for input_path in input_paths:
        if not input_path.exists():
            print(f"Skipping missing file: {input_path}")
            continue

        stem = input_path.stem
        out_name = f"{re.sub(r'\s*RAW$', '', stem, flags=re.I)} WORKSHEET.xlsx"
        output_path = output_dir / out_name
        print(f"Converting {input_path.name} -> {output_path.name}")
        build_workbook(input_path, output_path)

    print("Conversion complete.")


if __name__ == "__main__":
    main()
